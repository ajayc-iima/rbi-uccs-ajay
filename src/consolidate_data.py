import os
import re
import pandas as pd
import openpyxl
from datetime import datetime

# ---------------------------------------------------
# FILE AND DIRECTORY PATHS
# ---------------------------------------------------
try:
    # Assumes the script is in a 'src' folder, so goes one level up.
    BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
except NameError:
    # Fallback for when __file__ is not defined (e.g., in an interactive session)
    BASE_DIR = os.getcwd()

DATA_DIR = os.path.join(BASE_DIR, "data")
INTERIM_DIR = os.path.join(DATA_DIR, "interim")
PROCESSED_DIR = os.path.join(DATA_DIR, "processed")

# ---------------------------------------------------
# HELPER FUNCTIONS
# ---------------------------------------------------

def find_excel_files(directory):
    """
    Recursively finds all Excel (.xlsx) files in a given directory.

    Args:
        directory (str): The path to the directory to search.

    Returns:
        list: A list of full paths to the found Excel files.
    """
    excel_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.xlsx'):
                excel_files.append(os.path.join(root, file))
    return excel_files

def parse_survey_round(date_str):
    """
    Standardizes a date string from 'Mon-YY' format to 'YYYY-MM-01'.

    Args:
        date_str (str): The date string to parse (e.g., 'May-25').

    Returns:
        str: The formatted date string, or None if parsing fails.
    """
    if not isinstance(date_str, str):
        return None
    try:
        # The datetime object for the first day of the given month and year.
        dt_object = datetime.strptime(date_str, '%b-%y')
        return dt_object.strftime('%Y-%m-%d')
    except (ValueError, TypeError):
        return None

def extract_perception_category(title_from_cell):
    """
    Extracts a clean perception category from the title string found in the sheet.
    Example: 'Table 1: Perceptions... on Economic Situation' -> 'Economic Situation'
    """
    # Use regex to find the text after 'on ' or 'for '
    match = re.search(r'on\s(.+)|for\s(.+)', title_from_cell, re.IGNORECASE)
    if match:
        # return the first non-None group, stripped of extra chars
        category = next(g for g in match.groups() if g is not None)
        # Further clean up potential trailing characters like '*'
        return re.sub(r'[\*:]', '', category).strip()
    return "General" # Fallback category

def process_sheet(df, sheet_title):
    """
    Transforms a single sheet's DataFrame from wide to long format.

    Args:
        df (pd.DataFrame): The DataFrame from a single Excel sheet.
        sheet_title (str): The title of the table, used to derive the perception category.

    Returns:
        pd.DataFrame: A cleaned and transformed DataFrame in long format.
    """
    # Ensure the first column is named 'survey_round'
    if 'Survey Round' in df.columns:
        df = df.rename(columns={'Survey Round': 'survey_round'})
    else:
        # If the first column isn't the survey round, we can't process it.
        return pd.DataFrame()

    # Unpivot the table from wide to long format
    long_df = pd.melt(
        df,
        id_vars='survey_round',
        var_name='metric',
        value_name='response_percentage'
    )

    # --- Split the 'metric' column into 'perception_type' and 'response_category' ---
    
    # Define perception types
    perception_types = ['Current Perception', 'One year ahead Expectation']
    
    # Extract perception type
    long_df['perception_type'] = long_df['metric'].apply(
        lambda x: next((pt for pt in perception_types if str(x).startswith(pt)), 'Net Response')
    )
    
    # Extract response category by removing the perception type part
    long_df['response_category'] = long_df.apply(
        lambda row: str(row['metric']).replace(row['perception_type'], '').replace('-', '').strip(),
        axis=1
    )
    # Correct the response category for 'Net Response' rows
    long_df.loc[long_df['perception_type'] == 'Net Response', 'response_category'] = 'Net Response'

    # --- Clean and add remaining columns ---
    
    # Add the perception category from the sheet title
    long_df['perception_category'] = extract_perception_category(sheet_title)
    
    # Standardize the survey round date
    long_df['survey_round'] = long_df['survey_round'].apply(parse_survey_round)
    
    # --- Data Validation and Final Cleanup ---
    
    # Drop rows where essential data is missing
    long_df.dropna(subset=['survey_round', 'response_percentage'], inplace=True)
    
    # Convert percentage to a numeric type, coercing errors to NaN
    long_df['response_percentage'] = pd.to_numeric(long_df['response_percentage'], errors='coerce')
    long_df.dropna(subset=['response_percentage'], inplace=True)
    
    # Select and reorder final columns
    final_df = long_df[[
        'survey_round',
        'perception_category',
        'perception_type',
        'response_category',
        'response_percentage'
    ]]
    
    return final_df

# ---------------------------------------------------
# MAIN EXECUTION BLOCK
# ---------------------------------------------------
def main():
    """
    Main pipeline to read, process, consolidate, and save the survey data.
    """
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    
    excel_files = find_excel_files(INTERIM_DIR)
    if not excel_files:
        print(f"❌ No Excel files found in '{INTERIM_DIR}'. Please run the first script to generate them.")
        return

    print(f"✅ Found {len(excel_files)} Excel files to process.")
    
    all_processed_sheets = []
    
    for file_path in excel_files:
        print(f"  Processing '{os.path.basename(file_path)}'...")
        try:
            # Use openpyxl to manually read the file to bypass pandas parsing errors
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            for sheet_name in workbook.sheetnames:
                # Skip summary sheets and the table of contents sheet
                if sheet_name.lower().startswith('summary') or sheet_name.lower() == 'all table titles':
                    print(f"    Skipping non-data sheet: '{sheet_name}'")
                    continue

                sheet = workbook[sheet_name]
                
                # Convert sheet to a list of lists, containing only cell values
                data = []
                for row in sheet.iter_rows():
                    data.append([cell.value for cell in row])

                # Get category title from the first line of the sheet
                title_from_cell = sheet_name # Use sheet_name as a fallback
                if data and data[0] and isinstance(data[0][0], str):
                    title_from_cell = data[0][0]

                # Dynamically find the header row by looking for 'Survey Round'
                header_row_index = None
                for i, row_data in enumerate(data):
                    if row_data and 'Survey Round' in str(row_data[0]):
                        header_row_index = i
                        break
                
                if header_row_index is None:
                    print(f"    ⚠️ Could not find 'Survey Round' header in sheet '{sheet_name}'. Skipping.")
                    continue

                # Create DataFrame manually from the extracted data
                header = data[header_row_index]
                data_rows = data[header_row_index + 1:]
                
                df = pd.DataFrame(data_rows, columns=header)
                
                # Clean up the DataFrame
                df.dropna(how='all', inplace=True) # Drop rows that are completely empty
                df.dropna(axis=1, how='all', inplace=True) # Drop columns that are completely empty
                
                # Process the sheet, passing the title found in the cell
                processed_df = process_sheet(df, title_from_cell)
                if not processed_df.empty:
                    all_processed_sheets.append(processed_df)

        except Exception as e:
            # Catch any exception during file processing and report it
            print(f"    ⚠️ Could not process file '{os.path.basename(file_path)}'. Reason: {e}")

    if not all_processed_sheets:
        print("❌ No data could be consolidated. Exiting.")
        return
        
    # Combine all processed DataFrames into a single one
    consolidated_df = pd.concat(all_processed_sheets, ignore_index=True)
    
    # Final validation: drop any full duplicate rows
    consolidated_df.drop_duplicates(inplace=True)
    
    # Save the final dataset
    output_filename = "consolidated_uccs_data.csv"
    output_path = os.path.join(PROCESSED_DIR, output_filename)
    consolidated_df.to_csv(output_path, index=False)
    
    print("-" * 60)
    print(f"✨ All processing complete.")
    print(f"✅ Consolidated data saved to: '{output_path}'")
    print(f"Total rows in final dataset: {len(consolidated_df)}")

if __name__ == "__main__":
    main()
